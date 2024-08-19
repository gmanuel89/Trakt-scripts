#####
# Author: Manuel Galli
# e-mail: gmanuel89@gmail.com / manuel.galli@revvity.com
# Updated date: 2022-10-07
#####

## Import libraries
import openpyxl

## Write spreadsheet to file
def remove_default_empty_sheet_from_workbook(spreadsheet_workbook: openpyxl.Workbook) -> openpyxl.Workbook:
    # Verify if there is a sheet named "Sheet"
    if 'Sheet' in spreadsheet_workbook.sheetnames:
        # Retrieve the "Sheet" content
        sheet_content = spreadsheet_workbook['Sheet']
        # Determine that it is empty
        if sheet_content.max_column == 1 and sheet_content.max_row == 1:
            # Retrieve the single cell value and determine that it's empty
            if sheet_content['A1'].value == '' or sheet_content['A1'].value is None:
                # Remove the sheet
                spreadsheet_workbook.remove(spreadsheet_workbook['Sheet'])
    # return
    return spreadsheet_workbook
