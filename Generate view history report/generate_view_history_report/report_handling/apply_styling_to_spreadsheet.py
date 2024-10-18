#####
# Author: Manuel Galli
# e-mail: gmanuel89@gmail.com / manuel.galli@revvity.com
# Updated date: 2024-10-18
#####

## Import libraries and functions
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

## Apply styling to spreadsheet
def apply_styling_to_spreadsheet(spreadsheet_sheet: Worksheet, freeze_header=True, freeze_first_column=False) -> Worksheet:
    # Make the header bold
    for col in range(spreadsheet_sheet.max_column):
        cell_letter = get_column_letter(col+1)
        spreadsheet_sheet[cell_letter+'1'].font = Font(bold=True)
    # Autoresize columns
    for col_index, col in enumerate(spreadsheet_sheet.columns, 1):
        spreadsheet_sheet.column_dimensions[get_column_letter(col_index)].auto_size = True
    """
    column_letters = tuple(get_column_letter(col_number + 1) for col_number in range(spreadsheet_sheet.max_column))
    for column_letter in column_letters:
            spreadsheet_sheet.column_dimensions[column_letter].bestFit = True
    """
    # Freeze the header
    if freeze_header and not freeze_first_column:
        spreadsheet_sheet.freeze_panes = 'A2'
    # Freeze first column
    elif freeze_first_column and not freeze_header:
        spreadsheet_sheet.freeze_panes = 'B1'
    # Freeze the header and the first column
    elif freeze_header and freeze_first_column:
        spreadsheet_sheet.freeze_panes = 'B2'
    # return
    return spreadsheet_sheet
