## Import libraries and functions
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

## Apply styling to spreadsheet
def apply_styling_to_spreadsheet(spreadsheet_sheet: Worksheet, freeze_header=True, freeze_first_column=False) -> Worksheet:
    # Make the header bold
    for col in range(spreadsheet_sheet.max_column):
        cell_letter = get_column_letter(col+1)
        spreadsheet_sheet[cell_letter+'1'].font = Font(bold=True)
    # Autoresize columns
    for col_index, col in enumerate(spreadsheet_sheet.columns, 1):
        spreadsheet_sheet.column_dimensions[get_column_letter(col_index)].auto_size = True
    """
    column_letters = tuple(get_column_letter(col_number + 1) for col_number in range(spreadsheet_sheet.max_column))
    for column_letter in column_letters:
            spreadsheet_sheet.column_dimensions[column_letter].bestFit = True
    """
    # Freeze the header
    if freeze_header and not freeze_first_column:
        spreadsheet_sheet.freeze_panes = 'A2'
    # Freeze first column
    elif freeze_first_column and not freeze_header:
        spreadsheet_sheet.freeze_panes = 'B1'
    # Freeze the header and the first column
    elif freeze_header and freeze_first_column:
        spreadsheet_sheet.freeze_panes = 'B2'
    # return
    return spreadsheet_sheet

## Write spreadsheet to file
def remove_default_empty_sheet_from_workbook(spreadsheet_workbook: openpyxl.Workbook) -> openpyxl.Workbook:
    # Verify if there is a sheet named "Sheet"
    if 'Sheet' in spreadsheet_workbook.sheetnames:
        # Retrieve the "Sheet" content
        sheet_content = spreadsheet_workbook['Sheet']
        # Determine that it is empty
        if sheet_content.max_column == 1 and sheet_content.max_row == 1:
            # Retrieve the single cell value and determine that it's empty
            if sheet_content['A1'].value == '' or sheet_content['A1'].value is None:
                # Remove the sheet
                spreadsheet_workbook.remove(spreadsheet_workbook['Sheet'])
    # return
    return spreadsheet_workbook

## Write a CSV dictionary to a Spreadsheet
def write_spreadsheet_to_workbook(workbook_content: list[dict], sheet_name: str, spreadsheet_workbook: openpyxl.Workbook, freeze_header=True, freeze_first_column=False) -> openpyxl.Workbook:
    # Create a new Workbook if not provided
    if not spreadsheet_workbook:
        spreadsheet_workbook = openpyxl.Workbook()
    # Create dedicated sheet (truncate sheet name if it is too long)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[0:30]
    spreadsheet_workbook.create_sheet(sheet_name)
    spreadsheet_sheet = spreadsheet_workbook[sheet_name]
    # Write the content (using a numbered map)
    spreadsheet_column_map = {}
    # Header numbered map
    csv_header = list(workbook_content[0].keys())
    for i in range(len(csv_header)):
        spreadsheet_column_map[i+1] = csv_header[i]
    spreadsheet_sheet.append(spreadsheet_column_map)
    # Content (get the position number from the header map), for each row...
    for r in range(len(workbook_content)):
        row = workbook_content[r]
        row_dictionary_fixed_for_spreadsheet = row
        # Replace the key with the positional number (to write the cells in the spreadsheet)
        for i in range(len(csv_header)):
            row_dictionary_fixed_for_spreadsheet[i+1] = row.pop(csv_header[i])
        spreadsheet_sheet.append(row_dictionary_fixed_for_spreadsheet)
    # Apply styling to worksheet
    spreadsheet_sheet = apply_styling_to_spreadsheet(spreadsheet_sheet, freeze_header, freeze_first_column)
    # Print console message
    print('Writing spreadsheet: ' + str(sheet_name))
    # Return the same Workbook as input but with the added sheet
    return spreadsheet_workbook

## Write spreadsheet to file
def write_workbook(spreadsheet_workbook: openpyxl.Workbook, output_folder='', file_name='', remove_default_empty_sheet=True) -> None:
    # Remove the default empty sheet
    if remove_default_empty_sheet:
        spreadsheet_workbook = remove_default_empty_sheet_from_workbook(spreadsheet_workbook)
    # Fix the filename
    if file_name is None or file_name == '':
        file_name = 'Spreadsheet'
    if not file_name.endswith('.xlsx'):
        file_name = file_name + '.xlsx'
    if output_folder is not None and output_folder != '':
        file_name = output_folder + '/' + file_name
    # Save the file (not if empty)
    if len(spreadsheet_workbook.worksheets) > 0:
        spreadsheet_workbook.save(file_name)
